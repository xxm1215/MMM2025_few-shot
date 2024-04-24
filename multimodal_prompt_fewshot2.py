import glob, json, random, re, nltk, csv, argparse
import numpy as np
from openprompt.data_utils import InputExample, FewShotSampler
from openprompt.prompts import MixedTemplate, SoftTemplate, ManualTemplate
from openprompt.plms import load_plm
from openprompt import PromptDataLoader, PromptForClassification
from OpenPrompt.openprompt.prompts import SoftVerbalizer, ManualVerbalizer, KnowledgeableVerbalizer
import torch
import torch.nn.functional as F
from transformers import AdamW, get_linear_schedule_with_warmup
from sklearn.metrics import classification_report
from sklearn.model_selection import train_test_split
import clip
from PIL import Image
from sklearn.manifold import TSNE
import matplotlib.pyplot as plt
from sklearn.metrics import f1_score
import textblob
import os
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from sklearn.metrics import f1_score
from tqdm import tqdm
from lime import lime_image
from lime.lime_text import LimeTextExplainer



device = "cuda" if torch.cuda.is_available() else "cpu"


def xrange(x):
    return iter(range(x))

def test_set(all, train, dev, few_shot=True, if_dev=True):
    test = []
    if few_shot:
        used_id = []
        for idx in range(len(train)):
            used_id.append(train[idx].guid)
        if if_dev:
            for idx_1 in range(len(dev)):
                used_id.append(dev[idx_1].guid)

        for idx in range(len(all)):
            if all[idx].guid not in used_id:
                test.append(all[idx])
            else:
                continue
    return test

class Alpha(torch.nn.Module):
    def __init__(self):
        super().__init__()
        self.raw_beta = torch.nn.Parameter(data=torch.Tensor(0), requires_grad=True)

    def forward(self):  # no inputs
        beta = torch.sigmoid(self.raw_beta)  # get (0,1) value
        return beta

#--------------yj特征融合---------------
class Proj_layers(torch.nn.Module):
    def __init__(self):
        super().__init__()
        self.proj1 = torch.nn.Linear(1024, 768, device=device)
        self.ln1 = torch.nn.LayerNorm(768, device=device)
        self.proj2 = torch.nn.Linear(768, 768, device=device)
        self.ln2 = torch.nn.LayerNorm(768, device=device)

    def forward(self, txt, img):
        out_emb = torch.cat((img, txt), 1)#[1,1024]
        out_emb = self.ln1(F.relu(self.proj1(out_emb.float())))#[1,768]
        out_emb = self.ln2(F.relu(self.proj2(out_emb)))#[1,768]
        return out_emb
#------------------Cross-Attention---------
class CrossAttention(torch.nn.Module):
    def __init__(self, in_dim1, in_dim2, k_dim, v_dim, num_heads):
        super(CrossAttention, self).__init__()
        self.num_heads = num_heads
        self.k_dim = k_dim
        self.v_dim = v_dim
        self.proj_q1 = torch.nn.Linear(in_dim1, k_dim * num_heads, bias=False,device=device).half()
        self.proj_k2 = torch.nn.Linear(in_dim2, k_dim * num_heads, bias=False,device=device).half()
        self.proj_v2 = torch.nn.Linear(in_dim2, v_dim * num_heads, bias=False,device=device).half()
        self.proj_o = torch.nn.Linear(v_dim * num_heads, 512,device=device).half()#[1,512]
        
    def forward(self, x1, x2, mask=None):
        seq_len1, in_dim1 = 1,x1.size()#1*512
        seq_len2 = x2.size(0)

        q1 = self.proj_q1(x1).view(seq_len1, self.num_heads, self.k_dim).permute(1, 0, 2)
        k2 = self.proj_k2(x2).view(seq_len2, self.num_heads, self.k_dim).permute(1, 2, 0)
        v2 = self.proj_v2(x2).view(seq_len2, self.num_heads, self.v_dim).permute(1, 0 ,2)
        
        attn = torch.matmul(q1, k2) / self.k_dim**0.5
        
        if mask is not None:
            attn = attn.masked_fill(mask == 0, -1e9)
        
        attn = F.softmax(attn, dim=-1)
        output = torch.matmul(attn, v2).permute(1, 0, 2).contiguous().view(seq_len1, -1)
        output = self.proj_o(output)
        #print('cat_output_shape',output.shape)
        return output
#----------------the right part of model-------------
# def mini_batching(inputs):
#     mini_batch = []
#     sim_all = []
#     proj = Proj_layers()
#     cat=CrossAttention(512,512,512,512,1)#cross-attetion
#     sim = torch.nn.CosineSimilarity(dim=1, eps=1e-6)
#     sig = torch.nn.Sigmoid()
#     for item in inputs['guid']:
#         for sample in all_data:
#             if item == sample['id']:
#                 i_input = preprocess(Image.open(image_path + item + ".jpg")).unsqueeze(0).to(device)
#                 t_input = clip.tokenize(sample['txt'], truncate=True).to(device)

#                 i_emb = model.encode_image(i_input)#[1,512]
#                 #print('i_emb:',i_emb.size())
#                 t_emb = model.encode_text(t_input)#[1,512]
#                 #out_emb = proj(t_emb, i_emb)#特征融合拼接
#                 #dim1 query dim2
#                 #out_emb = cat(i_emb,t_emb)#特征融合cross-attetion
#                 out_embi_t=cat(i_emb,t_emb)
#                 out_embt_i=cat(t_emb,i_emb)
                
#                 out_emb_txt=out_embi_t+t_emb#残差连接[1,512]
#                 out_emb_img=out_embt_i+i_emb#残差连接[1,512]
                

                
#                 out_emb = proj(out_emb_txt, out_emb_img)#特征融合拼接[1,512]->[1,768]
                

#                 sim_all.append(sim(t_emb, i_emb))

#                 mini_batch.append(out_emb)
#     sim_all = torch.stack(sim_all).to(device).squeeze()
#     mini_batch = torch.stack(mini_batch).to(device).squeeze()
#     sim_mean = torch.mean(sim_all)
#     sim_std = torch.std(sim_all)
#     normalized_mini = sig((mini_batch - sim_mean) / sim_std)#计算相似度
#     mini_batch = normalized_mini * mini_batch#final fused feature

#     return mini_batch

def mini_batching(inputs):
    mini_batch = []
    sim_all = []
    proj = Proj_layers()
    cat=CrossAttention(512,512,512,512,4)#cross-attetion
    sim = torch.nn.CosineSimilarity(dim=1, eps=1e-6)
    sig = torch.nn.Sigmoid()
    for item in inputs['guid']:
        for sample in all_data:
            if item == sample['id']:
                i_input = preprocess(Image.open(image_path + item + ".jpg")).unsqueeze(0).to(device)
                t_input = clip.tokenize(sample['txt'], truncate=True).to(device)
                emo_score=sample['emo_socre']
                if emo_score==0:
                    emo_score=1e-2
                #normalized_polarity=sample['normalized_polarity']
                #normalized_polarity=torch.tensor(normalized_polarity)
                #情感分析
                #blob = textblob.TextBlob(sample['txt'])
                #情感results
                #result_sentiment = blob.sentiment

                # polarity=abs(result_sentiment.polarity)#情感极性
                #subjectivity=result_sentiment.subjectivity#主观
                # emo=polarity+subjectivity
                #emo_tensor = torch.tensor(emo)
                
                i_emb = model.encode_image(i_input).to(device)#[1,512]
                
                t_emb = (model.encode_text(t_input)*emo_score).to(device)#[1,512]
                out_embi_t=cat(i_emb,t_emb).to(device)
                out_embt_i=cat(t_emb,i_emb).to(device)
                
                out_emb_txt=out_embi_t.to(device)+t_emb.to(device)#残差连接[1,512]
                out_emb_img=out_embt_i.to(device)+i_emb.to(device)#残差连接[1,512]
                

                out_emb = proj(out_emb_txt, out_emb_img).to(device)#特征融合拼接[1,512]->[1,768]
                

                sim_all.append(sim(t_emb, i_emb))

                mini_batch.append(out_emb)
    sim_all = torch.stack(sim_all).to(device).squeeze()
    mini_batch = torch.stack(mini_batch).to(device).squeeze()
    sim_mean = torch.mean(sim_all).to(device)
    sim_std = torch.std(sim_all).to(device)
    normalized_mini = (sig((mini_batch - sim_mean) / sim_std))#计算相似度
    mini_batch = (normalized_mini * mini_batch).to(device)#final fused feature

    return mini_batch
###########fused3 Modality_Interation########
def modality_interation(inputs):
    inter_batch = []
    sim_all = []
    proj = Proj_layers()
    sim = torch.nn.CosineSimilarity(dim=1, eps=1e-6)
    sig = torch.nn.Sigmoid()
    for item in inputs['guid']:
        for sample in all_data:
            if item == sample['id']:
                i_input = preprocess(Image.open(image_path + item + ".jpg")).unsqueeze(0).to(device)
                t_input = clip.tokenize(sample['txt'], truncate=True).to(device)
                emo_score=sample['emo_socre']
                if emo_score==0:
                    emo_score=1e-2
                #normalized_polarity=sample['normalized_polarity']
                #normalized_polarity=torch.tensor(normalized_polarity)
                 #情感分析
                #blob = textblob.TextBlob(sample['txt'])
                #情感results
                #result_sentiment = blob.sentiment

                # polarity=abs(result_sentiment.polarity)#情感极性
                #subjectivity=result_sentiment.subjectivity#主观
                # emo=polarity+subjectivity
                
                i_emb = model.encode_image(i_input)#[1,512]
                t_emb = model.encode_text(t_input)#[1,512]
                
                out_emb = proj(t_emb, i_emb)#原始图文特征proj用于模态交互影响
                
                sim_all.append(sim(t_emb, i_emb))
                inter_batch.append(out_emb)
    sim_all = torch.stack(sim_all).to(device).squeeze()
    inter_batch = torch.stack(inter_batch).to(device).squeeze()
    sim_mean = torch.mean(sim_all)
    sim_std = torch.std(sim_all)
    normalized_mini = sig((inter_batch - sim_mean) / sim_std)#计算相似度
    inter_batch = normalized_mini * inter_batch

    return inter_batch
#333333333333333333333333333333
def model_predict(inputs):
    for item in inputs['guid']:
        for sample in all_data:
            if item == sample['id']:
                # i_input = preprocess(Image.open(image_path_goss + item + ".jpg")).unsqueeze(0).to(device)
                t_input = clip.tokenize(sample['txt'], truncate=True).to(device)
                outputs = model.encode_text(t_input)
                logits = model.verbalizer.process_outputs(outputs=outputs, batch=None)
                probabilities = torch.nn.functional.softmax(logits, dim=1).detach().cpu().numpy()
    return probabilities



def model_predict_images(inputs):

    for item in inputs['guid']:
        for sample in all_data:
            if item == sample['id']:
                i_input = preprocess(Image.open(image_path + item + ".jpg")).unsqueeze(0).to(device)
                outputs = model.encode_image(i_input)
                logits = model.verbalizer.process_outputs(outputs=outputs, batch=None)
                probabilities = torch.nn.functional.softmax(logits, dim=1).detach().cpu().numpy()
    return probabilities

    
    # # Convert PIL images to tensor and preprocess
    # inputs = torch.stack([preprocess(img).unsqueeze(0).to(device) for img in images])
    # outputs = model.encode_image(inputs)
    # logits = model.verbalizer.process_outputs(outputs=outputs, batch=None)
    # probabilities = torch.nn.functional.softmax(logits, dim=1).detach().cpu().numpy()
    # return probabilities

############
def start_training(model, train_dataloader, val_dataloader,
          test_dataloader, loss_function, optimizer, alpha, epoch):

    print("alpha is {}".format(alpha))
    saved_model = None
    val_f1_macro_in_alpha = 0
    val_loss_in_alpha = 20
    tolerant = 5
    saved_epoch = 0
    for epoch in range(epoch):
        tot_loss = 0
        print("===========EPOCH:{}=============".format(epoch+1))
        for inputs in tqdm(train_dataloader):
            if use_cuda:
                inputs = inputs.cuda()
                
            out = model.forward_without_verbalize(inputs)
            mini_batch = mini_batching(inputs)
            
            inter_batch=modality_interation(inputs)
            
            out = out+alpha*inter_batch
            
            logits, draw = model.verbalizer.process_outputs(outputs=out, batch=inputs)

            labels = inputs['label']
            loss = loss_function(logits, labels)
            loss.backward()
            tot_loss += loss.item()
            optimizer.step()
            optimizer.zero_grad()

        model.eval()

        allpreds = []
        alllabels = []
        eval_total_loss = 0
        dev_total_loss = 0
        c = 0
        val_step = 0
        for inputs in tqdm(val_dataloader):
            if use_cuda:
                inputs = inputs.cuda()
                
            out = model.forward_without_verbalize(inputs) # [batch_size, seq_len, feature]
            ####################out没有输出

            mini_batch = mini_batching(inputs)
            #out = alpha * out + mini_batch
            inter_batch=modality_interation(inputs)
            out = out+alpha*inter_batch
            # print("*" * 50)
            # print(out)
            # print(out.shape)
            # print("=" * 50)
            logits, draw = model.verbalizer.process_outputs(outputs=out, batch=inputs)
            
            # print("*" * 50)
            # print(logits)
            # print(logits.shape)
            # print("=" * 50)
            labels = inputs['label']
            eval_loss = loss_function(logits, labels)
            eval_total_loss += eval_loss.item()
            alllabels.extend(labels.cpu().tolist())
            allpreds.extend(torch.argmax(logits, dim=-1).cpu().tolist())
            dev_loss = eval_total_loss/(val_step+1
                                        )
            dev_total_loss += dev_loss
            c += 1

        dev_total_loss = dev_total_loss/(c+1)
        print("===========val_loss===========: ", dev_total_loss)
        acc = sum([int(i==j) for i,j in zip(allpreds, alllabels)])/len(allpreds)
        print("validation:",  acc)
        report_val = classification_report(alllabels, allpreds, output_dict=True,
                                           labels=[0,1], target_names=["real", "fake"],
                                           )
        f1_fake = report_val['fake']['f1-score']
        f1_real = report_val['real']['f1-score']
        f1_macro = report_val['macro avg']['f1-score']
        if float(val_loss_in_alpha) > dev_total_loss:
            # val_f1_macro_in_alpha = float(f1_macro)
            val_loss_in_alpha = dev_total_loss
            saved_model = model
            saved_epoch = epoch
            print("saving model at {} alpha with {} val_loss at Epoch {}".format(alpha, val_loss_in_alpha, saved_epoch+1))
        if epoch - saved_epoch >= tolerant:
            print("Early stopping at epoch {}.".format(epoch+1))
            break
    

    allpreds = []
    alllabels = []
    alllogits = []
    for inputs in tqdm(test_dataloader):
        if use_cuda:
            inputs = inputs.cuda()
        out = saved_model.forward_without_verbalize(inputs)
        mini_batch = mini_batching(inputs)
        #out = alpha * out + mini_batch
        inter_batch=modality_interation(inputs)
        out = out+alpha*inter_batch
        # alllogits.extend(out.detach().cpu().numpy())

        logits, draw = saved_model.verbalizer.process_outputs(outputs=out, batch=inputs)

        alllogits.extend(draw.detach().cpu().numpy())
        labels = inputs['label']
        alllabels.extend(labels.cpu().tolist())
        allpreds.extend(torch.argmax(logits, dim=-1).cpu().tolist())
    


  
    
    #保存到excel
    if str(args.data) == "goss":
        if t_or_f(args.full):
            directory = "./results_excelstd_sample/goss/full/"
        else:
            directory = "./results_excelstd_sample/goss/few/"
    else:
        if t_or_f(args.full):
            directory = "./results_excelstd_sample/poli/full/"
        else:
            directory = "./results_excelstd_sample/poli/few/"

    if not os.path.exists(directory):
        os.makedirs(directory)
    if t_or_f(args.full):
        file_path = "{}{}_{}al.xlsx".format(directory, args.data, args.alpha)
    else:
        file_path = "{}{}_{}st_{}al.xlsx".format(directory, args.data, args.shot+args.shot, args.alpha)
    if not os.path.exists(file_path):
        wb = Workbook()
        ws = wb.active
        # Define column headers
        ws.append(["seed", "acc", "macro_f1"])
    else:
        # Load existing workbook
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
    
        #计算acc与macro_f1
    acc = sum([int(i==j) for i,j in zip(allpreds, alllabels)])/len(allpreds)
    macro_f1 = f1_score(alllabels, allpreds, average='macro')
    row = (args.seed, acc, macro_f1)
    ws.append(row)
    wb.save(file_path)
    
    
    report_test = classification_report(alllabels, allpreds, labels=[0,1], target_names=["real", "fake"])

    if str(args.data) == "goss":
        if t_or_f(args.full):
            with open("./results/goss/full/{}_{}sd_{}al_{}.txt".format(args.data, args.seed,
                                                        args.alpha, args.full), "w") as out_file:
                print(report_test)
                # out_file.write(report_test)
        else:
            with open("./results/goss/few/{}_{}st_{}sd_{}al_{}.txt".format(args.data, args.shot+args.shot, args.seed,
                                                        args.alpha, args.full), "w") as out_file:
                print(report_test)
                # out_file.write(report_test)
    else:
        if t_or_f(args.full):
            with open("./results/poli/full/{}_{}sd_{}al_{}.txt".format(args.data, args.seed,
                                                        args.alpha, args.full), "w") as out_file:
                print(report_test)
                # out_file.write(report_test)
        else:
            with open("./results/poli/few/{}_{}st_{}sd_{}al_{}.txt".format(args.data, args.shot+args.shot, args.seed,
                                                        args.alpha, args.full), "w") as out_file:
                print(report_test)
                # out_file.write(report_test)

    return alllogits, allpreds


def tsne_plot(inputs, labels):
    "Creates and TSNE model and plots it"


    tsne_model = TSNE(perplexity=10, n_components=2, init='pca', n_iter=2500, random_state=23)
    new_values = tsne_model.fit_transform(inputs)

    x = []
    y = []
    for value in new_values:
        x.append(value[0])
        y.append(value[1])

    plt.figure(figsize=(16, 16))
    for i in range(len(labels)):
        if labels[i] == 0:
            plt.scatter(x[i], y[i], c='b')
        elif labels[i] == 1:
            plt.scatter(x[i], y[i], c='r')
        # plt.annotate(labels[i],
        #              xy=(x[i], y[i]),
        #              xytext=(5, 2),
        #              textcoords='offset points',
        #              ha='right',
        #              va='bottom')
    plt.show()



def t_or_f(arg):
    ua = str(arg).upper()
    if 'TRUE'.startswith(ua):
       return True
    elif 'FALSE'.startswith(ua):
       return False

def get_data(path3):
    all_data = []
    with open(path3, 'r') as inf:
        data = csv.reader(inf)
        next(data)
        for line in data:

            image_id = line[0]
            text = line[1]
            # normalized_polarity=line[3]
            label = line[2]  # 0-true, 1-fake
            emo_score=line[5]
            d = {}
            
            if len(text) > 0 and image_id + ".jpg" in [name.split("/")[-1] for name in image_files]:
                d["id"] = image_id
                d["txt"] = text
                d['label'] = int(label)
                d['emo_socre']=float(emo_score)
            else:
                continue
            all_data.append(d)

    return all_data


def get_n_trainable_params(model):
    # all trainable
    num_total_trainable = sum(p.numel() for p in model.parameters() if p.requires_grad)

    # split into the plm and classisifcation head
    num_plm_trainable = sum(p.numel() for p in model.plm.parameters() if p.requires_grad)

    # template trainable
    try:
        num_template_trainable = sum(p.numel() for p in model.template.soft_embedding.parameters() if p.requires_grad)
    except:
        num_template_trainable = 0

    # verbalizer trainable
    num_verbalizer_trainable = sum(p.numel() for p in model.verbalizer.parameters() if p.requires_grad)

    # assert sum of the two = total
    assert num_plm_trainable + num_template_trainable + num_verbalizer_trainable == num_total_trainable

    print(f"Number of trainable parameters of PLM: {num_plm_trainable}\n")
    print('#' * 50)
    print(f"Number of trainable parameters of template: {num_template_trainable}\n")
    print('#' * 50)
    print(f"Number of trainable parameters of verbalizer: {num_verbalizer_trainable}\n")
    print('#' * 50)
    print(f"Total number of trainable parameters of whole model: {num_total_trainable}")
    print(f"Verbalizer grouped_parameters_1 require_grad: {model.verbalizer.group_parameters_1[0].requires_grad}")


if __name__ == "__main__":

    parser = argparse.ArgumentParser(description="help")
    parser.add_argument("--alpha", type=float, help="alpha from 0-1")
    parser.add_argument("--seed",  type=int, help="seed from 1-5")
    parser.add_argument("--shot", type=int, help="shot from 1, 2, 4, 8, 50")
    parser.add_argument("--full", type=str, default=False)
    parser.add_argument("--data", type=str, default="goss")
    args = parser.parse_args()
    if t_or_f(args.full):
        print("<<<<<<<< Data:{}, Seed:{}, Alpha:{}, Full:{}".format(args.data, args.seed,
                                                                         args.alpha, args.full))
    else:
        print("<<<<<<<< Data:{}, Shot:{}, Seed:{}, Alpha:{}, Full:{}".format(args.data, args.shot+args.shot, args.seed,
                                                                         args.alpha, args.full))
    ## fakenewsnet data scripts
    #goss_path1 = ".../datasets/fakenewsnet/goss_img_all/*.jpg"
    goss_path1 = "/mnt/qust_521_big_2/20240330_CCL/dataset/fakenewsnet/goss_img_all/*.jpg"
    goss_path2 = "/mnt/qust_521_big_2/20240330_CCL/dataset/fakenewsnet/goss_img_all/"
    goss_path3 = '/mnt/qust_521_big_2/20240330_CCL/dataset/fakenewsnet/normalized_data_emoscore_goss.csv'

    poli_path1 = "/mnt/qust_521_big_2/20240330_CCL/dataset/fakenewsnet/poli_img_all/*.jpg"
    poli_path2 = "/mnt/qust_521_big_2/20240330_CCL/dataset/fakenewsnet/poli_img_all/"
    # poli_path3 = '/mnt/qust_521_big_2/20240330_CCL/dataset/fakenewsnet/politifact_multi1.csv'
    poli_path3 = "/mnt/qust_521_big_2/20240330_CCL/dataset/fakenewsnet/normalized_data_emoscore_poli.csv"
    if str(args.data) == "goss":
        image_files = glob.glob(goss_path1)
        image_path = goss_path2
        all_data = get_data(goss_path3)
    else:
        image_files = glob.glob(poli_path1)
        image_path = poli_path2
        all_data = get_data(poli_path3)


    model, preprocess = clip.load("ViT-B/32", device=device)

    # Initialize LIME Explainers here
    text_explainer = LimeTextExplainer(class_names=["real", "fake"])
    image_explainer = lime_image.LimeImageExplainer()

    dataset = []
    for idx, d in enumerate(all_data):
        input_example = InputExample(text_a=d['txt'], label=int(d['label']), guid=d['id'])
        dataset.append(input_example)

    if t_or_f(args.full):
        train, dev = train_test_split(dataset, test_size=0.2, shuffle=True)
        dev, test = train_test_split(dev, test_size=0.5, shuffle=True)

    else:
        sampler = FewShotSampler(num_examples_per_label=args.shot, num_examples_per_label_dev=args.shot, also_sample_dev=True)
        train, dev = sampler.__call__(train_dataset=dataset, seed=args.seed)

        test = test_set(dataset, train, dev, if_dev=True)

    pre_lm_dir = "/mnt/qust_521_big_2/public/roberta/roberta-base/"

    plm, tokenizer, model_config, WrapperClass = load_plm("roberta", pre_lm_dir)

    mytemplate = MixedTemplate(model=plm, tokenizer=tokenizer,
                               text='{"soft": None, "duplicate": 20}{"mask"}{"placeholder":"text_a"}')
    train_dataloader = PromptDataLoader(dataset=train, template=mytemplate, tokenizer=tokenizer,
                                        tokenizer_wrapper_class=WrapperClass, max_seq_length=512, decoder_max_length=3,
                                        batch_size=2, shuffle=True, teacher_forcing=False, predict_eos_token=False,
                                        truncate_method="tail",drop_last=True)

    validation_dataloader = PromptDataLoader(dataset=dev, template=mytemplate, tokenizer=tokenizer,
                                             tokenizer_wrapper_class=WrapperClass, max_seq_length=512,
                                             decoder_max_length=3,
                                             batch_size=2, shuffle=True, teacher_forcing=False, predict_eos_token=False,
                                             truncate_method="tail",drop_last=True)

    test_dataloader = PromptDataLoader(dataset=test, template=mytemplate, tokenizer=tokenizer,
                                       tokenizer_wrapper_class=WrapperClass, max_seq_length=512, decoder_max_length=3,
                                       batch_size=2, shuffle=True, teacher_forcing=False, predict_eos_token=False,
                                       truncate_method="tail",drop_last=True)

    myverbalizer = SoftVerbalizer(tokenizer, plm, num_classes=2)


    use_cuda = True
    prompt_model = PromptForClassification(plm=plm, template=mytemplate, verbalizer=myverbalizer, freeze_plm=False)
    if use_cuda:
        prompt_model = prompt_model.cuda()

    loss_func = torch.nn.CrossEntropyLoss()
    no_decay = ['bias', 'LayerNorm.weight']
    # it's always good practice to set no decay to biase and LayerNorm parameters
    optimizer_grouped_parameters1 = [
        {'params': [p for n, p in prompt_model.plm.named_parameters() if not any(nd in n for nd in no_decay)],
         'weight_decay': 0.01},
        {'params': [p for n, p in prompt_model.plm.named_parameters() if any(nd in n for nd in no_decay)],
         'weight_decay': 0.0}
    ]

    optimizer1 = AdamW(optimizer_grouped_parameters1, lr=3e-5)

    alpha = args.alpha
    get_n_trainable_params(prompt_model)

    all_logits, all_preds = start_training(model=prompt_model, train_dataloader=train_dataloader, val_dataloader=validation_dataloader,
          test_dataloader=test_dataloader, loss_function=loss_func, optimizer=optimizer1, alpha=alpha, epoch=20)

    get_n_trainable_params(prompt_model)
        # 在这里调用训练和评估模型的函数
# train_model(model, train_loader, validation_loader)
# evaluate_model(model, test_loader)

# LIME 可视化开始
# 确保模型, text_explainer, 和 image_explainer 已经初始化
    if 'text_explainer' in locals() and 'image_explainer' in locals():
        # For text
        idx = 10  # 确保您的数据集中至少有 idx+1 个样本
        text_instance = dataset[idx].text_a
        exp = text_explainer.explain_instance(text_instance, model_predict, num_features=6)
        # 将解释结果保存到 HTML 文件中或输出到控制台
        html_data = exp.as_html()
        with open("lime_result.html", "w") as file:
            file.write(html_data)

        # For image
        if idx < len(image_files):
            img_instance = Image.open(image_files[idx])  # 加载图片
            exp = image_explainer.explain_instance(np.array(img_instance), model_predict_images, top_labels=2, hide_color=0, num_samples=100)
            image, mask = exp.get_image_and_mask(exp.top_labels[0], positive_only=True, num_features=5, hide_rest=True)
            plt.imshow(image)
            plt.imshow(mask, cmap='jet', alpha=0.5)
            plt.title("LIME Explanation")
            plt.show()
        else:
            print("Index out of range for the image files")
    else:
        print("Make sure your model and explainers are properly initialized")
    # tsne_plot(np.array(all_logits), all_preds)
